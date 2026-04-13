import io
import os
from contextlib import asynccontextmanager

from dotenv import load_dotenv
load_dotenv()

from fastapi import FastAPI, Form, Request
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.templating import Jinja2Templates

from db import get_schema, get_tables_list, execute_query
from ai_helper import nl_to_sql

# ── app state ────────────────────────────────────────────────────────────────
_state: dict = {}


@asynccontextmanager
async def lifespan(app: FastAPI):
    _state["schema"]  = get_schema()
    _state["tables"]  = get_tables_list()
    yield


app = FastAPI(title="Gemio ERP NL→SQL", lifespan=lifespan)
templates = Jinja2Templates(directory="templates")


# ── routes ───────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    return templates.TemplateResponse(
        "index.html",
        {"request": request, "tables": _state.get("tables", [])},
    )


@app.post("/query")
async def query(question: str = Form(...)):
    schema_context = _state.get("schema", "")
    sql = nl_to_sql(question, schema_context)

    if sql.upper().startswith("ERROR:"):
        return JSONResponse({"error": sql, "sql": "", "columns": [], "rows": []})

    try:
        columns, rows = execute_query(sql)
        return JSONResponse({"error": None, "sql": sql, "columns": columns, "rows": rows})
    except Exception as e:
        return JSONResponse({"error": str(e), "sql": sql, "columns": [], "rows": []})


@app.post("/export-excel")
async def export_excel(request: Request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    data = await request.json()
    columns: list[str] = data.get("columns", [])
    rows: list[list[str]] = data.get("rows", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "查詢結果"

    # Header row
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    ws.append(columns)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row in rows:
        ws.append(row)

    # Auto-width
    for col_idx, _ in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(r, col_idx).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=query_result.xlsx"},
    )


# ── entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8080, reload=True)
